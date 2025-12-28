#!/usr/bin/env python3
"""
Script de importación de clientes desde Export.xlsx a la base de datos normalizada.
Importa: Clientes, Rutas, Pulperías, Usuarios, Cronograma de Visitas, Visitas Realizadas
"""

import openpyxl
import mysql.connector
from datetime import datetime
import os
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv('backend-laravel/.env')

# Configuración de base de datos
DB_CONFIG = {
    'host': os.getenv('DB_HOST', '127.0.0.1'),
    'database': os.getenv('DB_DATABASE', 'pandelpueblo'),
    'user': os.getenv('DB_USERNAME', 'root'),
    'password': os.getenv('DB_PASSWORD', 'Caracol@2024'),
    'port': int(os.getenv('DB_PORT', '3306'))
}

def connect_db():
    """Conectar a la base de datos MySQL"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG, autocommit=True)
        print(f"OK Conectado a la base de datos: {DB_CONFIG['database']}")
        return conn
    except Exception as e:
        print(f"ERROR conectando a la base de datos: {e}")
        raise

def get_or_create_usuario(cursor, nombre_encargado):
    """Obtener o crear usuario encargado"""
    if not nombre_encargado:
        return None

    # Buscar usuario existente
    cursor.execute("""
        SELECT id FROM users WHERE nombre LIKE %s LIMIT 1
    """, (nombre_encargado,))

    result = cursor.fetchone()
    if result:
        return result[0]

    # Crear nuevo usuario
    email = f"{nombre_encargado.lower().replace(' ', '')}@pandelpueblo.com"
    cursor.execute("""
        INSERT INTO users (nombre, correoElectronico, telefono, permiso, password, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        nombre_encargado.title(),
        email,
        '',
        'Vendedor',
        '$2y$10$92IXUNpkjO0rOQ5byMi.Ye4oKoEa3Ro9llC/.og/at2.uheWG/igi',  # password
        datetime.now(),
        datetime.now()
    ))

    usuario_id = cursor.lastrowid
    print(f"  -> Usuario creado: {nombre_encargado} (ID: {usuario_id})")
    return usuario_id

def get_or_create_ruta(cursor, nombre_ruta):
    """Obtener o crear ruta"""
    if not nombre_ruta or nombre_ruta == '0':
        return None

    # Limpiar nombre de ruta
    nombre_ruta = nombre_ruta.strip()

    # Buscar ruta existente
    cursor.execute("""
        SELECT id FROM rutas WHERE nombre LIKE %s LIMIT 1
    """, (nombre_ruta,))

    result = cursor.fetchone()
    if result:
        return result[0]

    # Crear nueva ruta
    cursor.execute("""
        INSERT INTO rutas (nombre, cantidadPulperias, cantidadClientes, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s)
        
    """, (
        nombre_ruta.title(),
        0,
        0,
        datetime.now(),
        datetime.now()
    ))

    ruta_id = cursor.lastrowid
    print(f"  -> Ruta creada: {nombre_ruta} (ID: {ruta_id})")
    return ruta_id

def get_or_create_pulperia(cursor, nombre_pulperia, ruta_id, orden):
    """Obtener o crear pulpería"""
    if not nombre_pulperia or nombre_pulperia == '0':
        return None

    # Limpiar nombre de pulpería
    nombre_pulperia = nombre_pulperia.strip()

    # Buscar pulpería existente
    cursor.execute("""
        SELECT id FROM pulperias WHERE nombre LIKE %s LIMIT 1
    """, (nombre_pulperia,))

    result = cursor.fetchone()
    if result:
        return result[0]

    # Crear nueva pulpería
    cursor.execute("""
        INSERT INTO pulperias (nombre, direccion, telefono, rutaId, orden, cantidadClientes, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        
    """, (
        nombre_pulperia.title(),
        '',
        '',
        ruta_id,
        orden if orden else 0,
        0,
        datetime.now(),
        datetime.now()
    ))

    pulperia_id = cursor.lastrowid
    print(f"  -> Pulpería creada: {nombre_pulperia} (ID: {pulperia_id})")
    return pulperia_id

def parse_gps(ubicacion_str):
    """Parsear string de ubicación GPS"""
    if not ubicacion_str:
        return None, None

    try:
        parts = str(ubicacion_str).split(',')
        if len(parts) == 2:
            lat = float(parts[0].strip())
            lon = float(parts[1].strip())
            return lat, lon
    except:
        pass

    return None, None

def import_cliente(cursor, row_data, encargados_map, rutas_map, pulperias_map):
    """Importar un cliente individual"""
    nombre = row_data.get('nombre', '').strip()
    if not nombre:
        return None

    direccion = row_data.get('direccion', '').strip() or ''
    telefono = str(row_data.get('telefono', '')).strip() or '0'
    ubicacion = row_data.get('ubicacion', '')
    pulperia_nombre = row_data.get('Pulperia', '')
    ruta_nombre = row_data.get('Ruta', '')
    orden = row_data.get('orden', None)
    encargado_nombre = row_data.get('encargado', '')
    dia = row_data.get('dia', '')
    dia2 = row_data.get('dia2', '')
    dia3 = row_data.get('dia3', '')
    fecha_str = row_data.get('fecha', '')

    # Parsear GPS
    latitude, longitude = parse_gps(ubicacion)

    # Obtener IDs de relaciones
    usuario_id = get_or_create_usuario(cursor, encargado_nombre) if encargado_nombre else None
    ruta_id = get_or_create_ruta(cursor, ruta_nombre) if ruta_nombre else None
    pulperia_id = get_or_create_pulperia(cursor, pulperia_nombre, ruta_id, orden) if pulperia_nombre else None

    # Insertar cliente
    cursor.execute("""
        INSERT INTO clientes (nombre, direccion, telefono, pulperiaId, latitude, longitude, usuarioId, orden, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        nombre,
        direccion,
        telefono,
        pulperia_id,
        latitude,
        longitude,
        usuario_id,
        orden if orden else None,
        datetime.now(),
        datetime.now()
    ))

    cliente_id = cursor.lastrowid

    # Importar cronograma de visitas (dia, dia2, dia3)
    dias_visita = []
    if dia and dia.strip():
        dias_visita.append(dia.strip().lower())
    if dia2 and dia2.strip():
        dias_visita.append(dia2.strip().lower())
    if dia3 and dia3.strip():
        dias_visita.append(dia3.strip().lower())

    for idx, dia_semana in enumerate(dias_visita):
        cursor.execute("""
            INSERT INTO cronograma_visitas (clienteId, dia_semana, orden, activo, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE activo=activo
        """, (
            cliente_id,
            dia_semana,
            idx + 1,
            True,
            datetime.now(),
            datetime.now()
        ))

    # Importar visita realizada si hay fecha
    if fecha_str:
        try:
            if isinstance(fecha_str, str):
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
            else:
                fecha = fecha_str

            cursor.execute("""
                INSERT INTO visitas_clientes (clienteId, fecha, realizada, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                cliente_id,
                fecha,
                True,  # Si tiene fecha, se asume que fue realizada
                datetime.now(),
                datetime.now()
            ))
        except:
            pass

    return cliente_id

def main():
    """Función principal de importación"""
    print("=" * 80)
    print("IMPORTACIÓN DE CLIENTES DESDE EXCEL")
    print("=" * 80)

    # Conectar a la base de datos
    conn = connect_db()
    cursor = conn.cursor()

    # Cargar Excel
    print("\nCargando archivo Export.xlsx...")
    wb = openpyxl.load_workbook('Export.xlsx')
    sheet = wb['A']

    # Leer encabezados
    headers = []
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            headers.append(header)

    print(f"OK Archivo cargado: {sheet.max_row - 1} clientes encontrados\n")

    # Mapas para cachear relaciones
    encargados_map = {}
    rutas_map = {}
    pulperias_map = {}

    # Procesar clientes
    clientes_importados = 0
    clientes_error = 0

    try:
        for row_num in range(2, sheet.max_row + 1):
            # Leer datos de la fila
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_num, column=col_idx).value
                row_data[header] = cell_value

            nombre = row_data.get('nombre', '')
            if not nombre:
                continue

            try:
                cliente_id = import_cliente(cursor, row_data, encargados_map, rutas_map, pulperias_map)
                if cliente_id:
                    clientes_importados += 1
                    if clientes_importados % 10 == 0:
                        print(f"  Procesados: {clientes_importados} clientes...")
            except Exception as e:
                clientes_error += 1
                import traceback
                print(f"  ERROR Error importando {nombre}: {e}")
                traceback.print_exc()
                if clientes_error >= 3:
                    print("\nDETENIENDO: Demasiados errores consecutivos")
                    break

        # Commit de la transacción
        conn.commit()

        print("\n" + "=" * 80)
        print("RESUMEN DE IMPORTACIÓN")
        print("=" * 80)
        print(f"OK Clientes importados exitosamente: {clientes_importados}")
        print(f"ERROR Clientes con error: {clientes_error}")
        print(f"Total procesados: {clientes_importados + clientes_error}")

        # Mostrar estadísticas
        cursor.execute("SELECT COUNT(*) FROM users WHERE permiso = 'Vendedor'")
        print(f"\nUsuarios (encargados) creados: {cursor.fetchone()[0]}")

        cursor.execute("SELECT COUNT(*) FROM rutas")
        print(f"Rutas creadas: {cursor.fetchone()[0]}")

        cursor.execute("SELECT COUNT(*) FROM pulperias")
        print(f"Pulperías creadas: {cursor.fetchone()[0]}")

        cursor.execute("SELECT COUNT(*) FROM clientes")
        print(f"Total clientes en BD: {cursor.fetchone()[0]}")

        cursor.execute("SELECT COUNT(*) FROM cronograma_visitas")
        print(f"Cronogramas de visita creados: {cursor.fetchone()[0]}")

        cursor.execute("SELECT COUNT(*) FROM visitas_clientes")
        print(f"Visitas registradas: {cursor.fetchone()[0]}")

        print("\nOK Importación completada exitosamente!")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR Error durante la importación: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == '__main__':
    main()
